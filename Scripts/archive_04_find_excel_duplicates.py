
import os, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Specifying archive and duplicate workbook paths
workbook_path = 'D:\Archive.xlsx'

# Creating dataframe from archive workbook and an empty dataframe with same columns 
df_xlsx = pd.read_excel(workbook_path)
df_duplicate = pd.DataFrame(columns=df_xlsx.columns)

# Generating list after stripping extensions from 'Content' column values
lst = [os.path.splitext(x)[0] for x in df_xlsx['Content']]

# Converting list values to lower case strings
for a in lst:
    a = str(a).lower()

# Checking repeated strings from list then adding duplicate 'Content' column rows from archive dataframe to empty dataframe using indexes
for i in range(0, len(lst)):
    count = 0
    # Ignoring 'Content' values where 'Directory' is equal to 'Music Videos' 
    if df_xlsx.at[i, 'Directory'] != 'Music Videos':
        for j in range(i+1, len(lst)):
            if(lst[i] == lst[j]):
                count = count + 1
                if count == 1:
                    df_duplicate.loc[df_xlsx.index[i]] = df_xlsx.iloc[i]
                df_duplicate.loc[df_xlsx.index[j]] = df_xlsx.iloc[j]

# Incrementing dataframe index by 2
df_duplicate.index = df_duplicate.index + 2

# Inserting Index column to dataframe
df_duplicate.insert(0, "Index", df_duplicate.index)

# Loading Archive workbook
wb = openpyxl.load_workbook(workbook_path)

# Recreating 'Duplicates' sheet in workbook
if 'Duplicates' in wb.sheetnames:
    wb.remove(wb['Duplicates'])
wb.create_sheet('Duplicates', 3)
ws = wb['Duplicates']

# Writing dataframe to excel sheet using dataframe_to_rows method
for r in dataframe_to_rows(df_duplicate, index=False, header=True):
    ws.append(r)

# Auto-adjusting width of columns
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length

# Saving Archive workbook
wb.save(filename = workbook_path)
