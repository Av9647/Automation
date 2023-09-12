
import os, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Declaring folder and file paths
archive_path = r'D:\Extra\Archive II'
workbook_path = 'D:\Archive.xlsx'
excel_entry = ['Category', 'Directory', 'Content', 'Size', 'Location']

# Loading Archive.xlsx using openpyxl module
wb = openpyxl.load_workbook(workbook_path)
ws = wb['Archive']

# Declaring function to calculate file size
def sizeof_fmt(num, suffix="B"):
    for unit in ["", " K", " M", " G", " T", " P", " E", " Z"]:
        if abs(num) < 1024.0:
            return f"{num:3.2f}{unit}{suffix}"
        num /= 1024.0
    return f"{num:.2f}Yi{suffix}"

# Declaring function to obtain cell data from rows
def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

# Assigning excel data to a list
existing_values = list(iter_rows(ws))

# Appending files in Archive II folder to excel object
for folder, subfolders, files in os.walk(archive_path):
    for file in files:
        string_1 = folder.split("Archive II\\",1)[1]
        category = string_1.split("\\",1)[0]
        directory = string_1.split("\\",1)[1]
        if 'Lossy' in directory:
            directory = directory.split("\\",2)[0] + "\\" + directory.split("\\",2)[2]
        content = os.path.join(folder, file)
        size = sizeof_fmt(os.stat(content).st_size)
        location = 'Elements I'
        excel_entry[0] = category
        excel_entry[1] = directory
        excel_entry[2] = file
        excel_entry[3] = size
        excel_entry[4] = location
        if excel_entry not in existing_values:
            ws.append(excel_entry)

# Assigning worksheet data to a dataframe then sorting rows alphabetically ignoring case
data = ws.values
columns = next(data)[0:]
df = pd.DataFrame(data, columns=columns)
sorted_values = df.sort_values(by=['Category', 'Directory', 'Content', 'Size', 'Location'], na_position='first', key=lambda col: col.str.lower())

# Removing existing data in Archive.xlsx
last_non_empty_index = len(list(ws.rows))
ws.delete_rows(2, last_non_empty_index)

# Appending sorted dataframe to Archive.xlsx
for r in dataframe_to_rows(sorted_values, index=False, header=False):
    ws.append(r)

# Saving changes to Archive.xlsx
wb.save(filename = workbook_path)
